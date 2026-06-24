#!/usr/bin/env python3
"""
Taruu financial model generator.
Emits financial-model.xlsx + prints key figures for FINANCIAL-MODEL.md.

Locked decisions (2026-06-24):
  - Participation vote: gross ₪3, platform skims 30% (₪0.90), treasury gets 70% (₪2.10).
  - Create vote: gross ₪50, 100% platform revenue (treasury funded by the 70% participation share).
  - Paddle = Merchant of Record: 5% + $0.50 / transaction.
  - Target: ₪30–45k/mo combined take-home (midpoint ₪37.5k) → ~₪50k/mo platform gross profit pre personal tax.
  - FB agents: human-in-the-loop only.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

USD_ILS = 3.7                # assumption, early-2026
PADDLE_PCT = 0.05
PADDLE_FIXED = 0.50 * USD_ILS  # ₪1.85
VOTE_GROSS = 3.0
SKIM = 0.30
CREATE_GROSS = 50.0
FIXED_COSTS = 1200.0         # monthly infra+tooling (₪)
TAKEHOME_TARGET = 37500.0    # combined net midpoint
EFF_TAX = 0.27               # effective income tax + Bituach Leumi for 2x osek murshe
GROSS_PROFIT_TARGET = round(TAKEHOME_TARGET / (1 - EFF_TAX) + FIXED_COSTS)  # ~₪52.5k

def paddle_fee(amount):
    return PADDLE_PCT * amount + PADDLE_FIXED

# ---- Unit economics: participation under single-charge vs wallet packs ----
def vote_pack(pack_ils, votes):
    fee = paddle_fee(pack_ils)
    skim_amt = SKIM * pack_ils
    platform_net = skim_amt - fee          # platform bears processing out of its skim
    treasury = (1 - SKIM) * pack_ils
    return {
        "pack": pack_ils, "votes": votes, "paddle_fee": round(fee, 2),
        "skim": round(skim_amt, 2), "platform_net": round(platform_net, 2),
        "platform_per_vote": round(platform_net / votes, 3),
        "treasury_total": round(treasury, 2), "treasury_per_vote": round(treasury / votes, 3),
    }

packs = [
    vote_pack(3, 1),     # single charge — the broken baseline
    vote_pack(30, 10),
    vote_pack(60, 20),
    vote_pack(150, 50),
    vote_pack(300, 100),
]

# create vote
create_fee = paddle_fee(CREATE_GROSS)
create_net = CREATE_GROSS - create_fee   # 100% platform minus processing

# ---- Scenarios ----
# Driver lens: the P&L is CREATE-led (each paid create nets ~₪45.65 vs ~₪0.7/participation).
# per_vote_net assumes the ₪60 pack blend for engaged voters.
PV_NET = vote_pack(60, 20)["platform_per_vote"]   # ~0.658

def scenario(name, ambassadors, creates_paid_mo, participants_per_create, extra_costs):
    # total creates = ambassador-funded(free) + paid. Funded creates still drive participation.
    funded_creates = ambassadors            # ~1 funded first-vote per active ambassador / ramp
    total_creates = funded_creates + creates_paid_mo
    participations = total_creates * participants_per_create
    rev_creates = creates_paid_mo * create_net
    rev_part = participations * PV_NET
    cac = funded_creates * 0.0              # we waive the ₪50 (comp, no Paddle) → ₪0 cash, opportunity only
    gross = rev_creates + rev_part
    costs = FIXED_COSTS + extra_costs
    gp = gross - costs
    takehome = max(0, gp) * (1 - EFF_TAX)
    return {
        "name": name, "ambassadors": ambassadors, "creates_paid": creates_paid_mo,
        "funded_creates": funded_creates, "total_creates": total_creates,
        "participants_per_create": participants_per_create, "participations": participations,
        "rev_creates": round(rev_creates), "rev_part": round(rev_part),
        "gross": round(gross), "costs": round(costs), "gross_profit": round(gp),
        "est_takehome": round(takehome),
    }

scenarios = [
    scenario("Break-even", 8, 25, 30, 200),
    scenario("Lean (ramp)", 20, 120, 40, 500),
    scenario("Target (living)", 50, 600, 50, 1200),
    scenario("Scale", 120, 1800, 70, 2500),
]

# creates-only break-even (cover costs, zero salary)
be_creates_only = (FIXED_COSTS) / create_net
# creates-only to hit gross-profit target
target_creates_only = (GROSS_PROFIT_TARGET) / create_net

# ---------- Build workbook ----------
wb = Workbook()
H = Font(bold=True, color="FFFFFF")
HEAD = PatternFill("solid", fgColor="14110E")
RED = PatternFill("solid", fgColor="E0301E")
SUB = Font(bold=True)
THIN = Border(*[Side(style="thin", color="DDDDDD")] * 4)
money = '#,##0'

def style_header(ws, row, ncols, fill=HEAD):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H; cell.fill = fill; cell.alignment = Alignment(horizontal="center")

def autosize(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, 10), 40)

# Sheet 1: Assumptions
ws = wb.active; ws.title = "Assumptions"
ws["A1"] = "TARUU — FINANCIAL MODEL"; ws["A1"].font = Font(bold=True, size=16)
ws["A2"] = "Locked 2026-06-24 · all figures ₪ unless noted · rate USD/ILS=%.2f" % USD_ILS
rows = [
    ("Input", "Value", "Note"),
    ("USD/ILS rate", USD_ILS, "assumption, early 2026"),
    ("Paddle fee %", "%.0f%%" % (PADDLE_PCT*100), "Merchant-of-Record"),
    ("Paddle fixed / txn", round(PADDLE_FIXED, 2), "$0.50"),
    ("Vote gross", VOTE_GROSS, "participation fee"),
    ("Platform skim", "%.0f%%" % (SKIM*100), "of each participation"),
    ("Treasury share", "%.0f%%" % ((1-SKIM)*100), "funds the civic decision"),
    ("Create gross", CREATE_GROSS, "100% to platform"),
    ("Fixed costs / mo", FIXED_COSTS, "infra + tooling"),
    ("Combined take-home target", TAKEHOME_TARGET, "midpoint of ₪30–45k"),
    ("Effective tax+BL", "%.0f%%" % (EFF_TAX*100), "2x osek murshe est."),
    ("→ Gross-profit target / mo", GROSS_PROFIT_TARGET, "to net the take-home"),
]
for i, r in enumerate(rows, start=4):
    for j, v in enumerate(r, start=1):
        ws.cell(row=i, column=j, value=v)
style_header(ws, 4, 3)
autosize(ws)

# Sheet 2: Unit economics
ws = wb.create_sheet("Unit Economics")
ws["A1"] = "PARTICIPATION — single charge vs wallet packs"; ws["A1"].font = Font(bold=True, size=13)
cols = ["Pack ₪", "Votes", "Paddle fee", "Skim 30%", "Platform NET", "Platform/vote", "Treasury", "Treasury/vote"]
for j, c in enumerate(cols, start=1): ws.cell(row=3, column=j, value=c)
style_header(ws, 3, len(cols))
for i, p in enumerate(packs, start=4):
    vals = [p["pack"], p["votes"], p["paddle_fee"], p["skim"], p["platform_net"],
            p["platform_per_vote"], p["treasury_total"], p["treasury_per_vote"]]
    for j, v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=j, value=v)
        if i == 4:  # single-charge row — flag red
            cell.fill = RED; cell.font = Font(color="FFFFFF", bold=True)
ws.cell(row=4+len(packs)+1, column=1, value="↑ ₪3 single charge loses money — platform NET negative. Wallet packs fix it.")
ws.cell(row=4+len(packs)+1, column=1).font = Font(italic=True, color="E0301E")
r0 = 4 + len(packs) + 3
ws.cell(row=r0, column=1, value="CREATE VOTE").font = SUB
for j, (k, v) in enumerate([("Gross", CREATE_GROSS), ("Paddle fee", round(create_fee,2)),
                            ("Platform NET", round(create_net,2)), ("Treasury", 0)], start=1):
    ws.cell(row=r0+1, column=j*1, value=k).font = SUB
    ws.cell(row=r0+2, column=j*1, value=v)
autosize(ws)

# Sheet 3: Scenarios
ws = wb.create_sheet("Scenarios")
ws["A1"] = "MONTHLY SCENARIOS (create-led P&L)"; ws["A1"].font = Font(bold=True, size=13)
cols = ["Scenario", "Ambassadors", "Paid creates/mo", "Funded creates", "Total creates",
        "Part./create", "Participations", "Rev creates", "Rev part.", "Gross rev",
        "Costs", "Gross profit", "Est. take-home"]
for j, c in enumerate(cols, start=1): ws.cell(row=3, column=j, value=c)
style_header(ws, 3, len(cols))
for i, s in enumerate(scenarios, start=4):
    vals = [s["name"], s["ambassadors"], s["creates_paid"], s["funded_creates"], s["total_creates"],
            s["participants_per_create"], s["participations"], s["rev_creates"], s["rev_part"],
            s["gross"], s["costs"], s["gross_profit"], s["est_takehome"]]
    for j, v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=j, value=v)
        if s["name"].startswith("Target"):
            cell.fill = PatternFill("solid", fgColor="FCE9C8")
autosize(ws)

# Sheet 4: Break-even
ws = wb.create_sheet("Break-even")
ws["A1"] = "BREAK-EVEN (creates-only lens)"; ws["A1"].font = Font(bold=True, size=13)
data = [
    ("Metric", "Value", "Meaning"),
    ("Create NET (after Paddle)", round(create_net, 2), "₪ per paid create"),
    ("Cover fixed costs", round(be_creates_only, 1), "paid creates/mo (zero salary)"),
    ("  = per day", round(be_creates_only/30, 1), "creates/day"),
    ("Hit gross-profit target", round(target_creates_only), "paid creates/mo for ₪%s GP" % GROSS_PROFIT_TARGET),
    ("  = per day", round(target_creates_only/30), "creates/day"),
    ("Participation-only alt.", round(GROSS_PROFIT_TARGET / PV_NET), "paid votes/mo if creates=0 (unrealistic)"),
]
for i, r in enumerate(data, start=3):
    for j, v in enumerate(r, start=1):
        ws.cell(row=i, column=j, value=v)
style_header(ws, 3, 3)
autosize(ws)

wb.save("growth/financial-model.xlsx")

# ---- print key figures for the markdown ----
print("=== UNIT ECONOMICS ===")
for p in packs:
    print(f"  pack ₪{p['pack']:>4} ({p['votes']:>3} votes): paddle ₪{p['paddle_fee']:>5} | "
          f"platform NET ₪{p['platform_net']:>7} ({p['platform_per_vote']:+.3f}/vote) | "
          f"treasury/vote ₪{p['treasury_per_vote']}")
print(f"  CREATE ₪50: paddle ₪{create_fee:.2f} -> platform NET ₪{create_net:.2f}")
print("=== BREAK-EVEN ===")
print(f"  cover costs: {be_creates_only:.1f} creates/mo ({be_creates_only/30:.1f}/day)")
print(f"  GP target ₪{GROSS_PROFIT_TARGET}: {target_creates_only:.0f} creates/mo ({target_creates_only/30:.0f}/day)")
print(f"  participation-only equiv: {GROSS_PROFIT_TARGET/PV_NET:,.0f} votes/mo (why creates lead)")
print("=== SCENARIOS ===")
for s in scenarios:
    print(f"  {s['name']:<16} amb={s['ambassadors']:>3} paidCreate={s['creates_paid']:>4} "
          f"part={s['participations']:>6} | gross ₪{s['gross']:>6} GP ₪{s['gross_profit']:>6} "
          f"takehome ₪{s['est_takehome']:>6}")
